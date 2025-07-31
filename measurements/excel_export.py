import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from django.http import HttpResponse
from django.utils import timezone
from django.db import models
from io import BytesIO
import logging

logger = logging.getLogger(__name__)


class GoatMeasurementExporter:
    """Utility class for exporting goat measurements to Excel"""
    
    def __init__(self):
        self.workbook = Workbook()
        
    def export_user_measurements(self, user, measurements_queryset=None):
        """
        Export all measurements for a user to Excel
        
        Args:
            user: Django User object
            measurements_queryset: Optional queryset to filter measurements
            
        Returns:
            HttpResponse with Excel file
        """
        try:
            if measurements_queryset is None:
                from .models import MorphometricMeasurement
                measurements_queryset = MorphometricMeasurement.objects.filter(
                    goat__owner=user
                ).order_by('-measurement_date')
            
            # Remove default sheet
            self.workbook.remove(self.workbook.active)
            
            # Create summary sheet
            self._create_summary_sheet(user, measurements_queryset)
            
            # Create detailed measurements sheet
            self._create_measurements_sheet(measurements_queryset)
            
            # Create goats overview sheet
            self._create_goats_sheet(user)
            
            # Create analysis sheet
            self._create_analysis_sheet(measurements_queryset)
            
            # Save to BytesIO
            excel_buffer = BytesIO()
            self.workbook.save(excel_buffer)
            excel_buffer.seek(0)
            
            # Create HTTP response
            response = HttpResponse(
                excel_buffer.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            
            filename = f"GoatMorpho_Measurements_{user.username}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            return response
            
        except Exception as e:
            logger.error(f"Error creating Excel export: {e}")
            raise
    
    def _create_summary_sheet(self, user, measurements_queryset):
        """Create summary sheet with user and measurement overview"""
        ws = self.workbook.create_sheet("Summary", 0)
        
        # Title
        ws['A1'] = "GoatMorpho - Measurement Report"
        ws['A1'].font = Font(size=16, bold=True, color="2F5F8F")
        ws.merge_cells('A1:E1')
        
        # User information
        ws['A3'] = "User Information"
        ws['A3'].font = Font(size=12, bold=True)
        
        ws['A4'] = "Name:"
        ws['B4'] = f"{user.first_name} {user.last_name}"
        ws['A5'] = "Username:"
        ws['B5'] = user.username
        ws['A6'] = "Email:"
        ws['B6'] = user.email
        ws['A7'] = "Member Since:"
        ws['B7'] = user.date_joined.strftime('%Y-%m-%d')
        
        # Statistics
        ws['A9'] = "Statistics"
        ws['A9'].font = Font(size=12, bold=True)
        
        total_measurements = measurements_queryset.count()
        unique_goats = measurements_queryset.values('goat').distinct().count()
        
        ws['A10'] = "Total Measurements:"
        ws['B10'] = total_measurements
        ws['A11'] = "Unique Goats:"
        ws['B11'] = unique_goats
        ws['A12'] = "Report Generated:"
        ws['B12'] = timezone.now().strftime('%Y-%m-%d %H:%M:%S')
        
        # Style the summary
        for row in range(1, 13):
            for col in range(1, 6):
                cell = ws.cell(row=row, column=col)
                if row in [1, 3, 9]:  # Headers
                    cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
    
    def _create_measurements_sheet(self, measurements_queryset):
        """Create detailed measurements sheet"""
        ws = self.workbook.create_sheet("Detailed Measurements")
        
        # Convert measurements to DataFrame
        data = []
        for measurement in measurements_queryset:
            row = {
                'Measurement ID': str(measurement.id),
                'Goat Name': measurement.goat.name or 'Unnamed',
                'Goat Breed': measurement.goat.breed or 'Unknown',
                'Measurement Date': measurement.measurement_date.strftime('%Y-%m-%d %H:%M:%S'),
                'Confidence Score': round(measurement.confidence_score, 3),
                'Reference Length (cm)': measurement.reference_object_length_cm or 'N/A',
                
                # Morphometric measurements
                'Height at Withers (cm)': round(measurement.hauteur_au_garrot, 2) if measurement.hauteur_au_garrot else 'N/A',
                'Body Length (cm)': round(measurement.body_length, 2) if measurement.body_length else 'N/A',
                'Chest Circumference (cm)': round(measurement.tour_de_poitrine, 2) if measurement.tour_de_poitrine else 'N/A',
                'Height at Croup (cm)': round(measurement.hauteur_au_sacrum, 2) if measurement.hauteur_au_sacrum else 'N/A',
                'Chest Width (cm)': round(measurement.largeur_poitrine, 2) if measurement.largeur_poitrine else 'N/A',
                'Hip Width (cm)': round(measurement.largeur_hanche, 2) if measurement.largeur_hanche else 'N/A',
                'Head Length (cm)': round(measurement.longueur_tete, 2) if measurement.longueur_tete else 'N/A',
                'Head Width (cm)': round(measurement.largeur_tete, 2) if measurement.largeur_tete else 'N/A',
                'Ear Length (cm)': round(measurement.longueur_oreille, 2) if measurement.longueur_oreille else 'N/A',
                'Neck Length (cm)': round(measurement.longueur_cou, 2) if measurement.longueur_cou else 'N/A',
                'Neck Circumference (cm)': round(measurement.tour_du_cou, 2) if measurement.tour_du_cou else 'N/A',
                'Tail Length (cm)': round(measurement.longueur_queue, 2) if measurement.longueur_queue else 'N/A',
            }
            data.append(row)
        
        if data:
            df = pd.DataFrame(data)
            
            # Write headers
            for col_num, column_title in enumerate(df.columns, 1):
                cell = ws.cell(row=1, column=col_num, value=column_title)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="2F5F8F", end_color="2F5F8F", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Write data
            for row_num, row_data in enumerate(df.values, 2):
                for col_num, value in enumerate(row_data, 1):
                    ws.cell(row=row_num, column=col_num, value=value)
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # Max width of 50
                ws.column_dimensions[column_letter].width = adjusted_width
        else:
            ws['A1'] = "No measurements found"
    
    def _create_goats_sheet(self, user):
        """Create goats overview sheet"""
        ws = self.workbook.create_sheet("Goats Overview")
        
        from .models import Goat, MorphometricMeasurement
        
        goats = Goat.objects.filter(owner=user)
        
        headers = ['Goat ID', 'Name', 'Breed', 'Age (months)', 'Sex', 'Weight (kg)', 'Total Measurements', 'Latest Measurement', 'Average Confidence']
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2F5F8F", end_color="2F5F8F", fill_type="solid")
        
        # Write goat data
        for row_num, goat in enumerate(goats, 2):
            measurements = MorphometricMeasurement.objects.filter(goat=goat)
            
            ws.cell(row=row_num, column=1, value=str(goat.id))
            ws.cell(row=row_num, column=2, value=goat.name or 'Unnamed')
            ws.cell(row=row_num, column=3, value=goat.breed or 'Unknown')
            ws.cell(row=row_num, column=4, value=goat.age_months if goat.age_months else 'Unknown')
            ws.cell(row=row_num, column=5, value=goat.get_sex_display() if goat.sex else 'Unknown')
            ws.cell(row=row_num, column=6, value=float(goat.weight_kg) if goat.weight_kg else 'Unknown')
            ws.cell(row=row_num, column=7, value=measurements.count())
            
            latest = measurements.order_by('-measurement_date').first()
            ws.cell(row=row_num, column=8, value=latest.measurement_date.strftime('%Y-%m-%d') if latest else 'N/A')
            
            avg_confidence = measurements.aggregate(avg_conf=models.Avg('confidence_score'))['avg_conf']
            ws.cell(row=row_num, column=9, value=round(avg_confidence, 3) if avg_confidence else 'N/A')
    
    def _create_analysis_sheet(self, measurements_queryset):
        """Create statistical analysis sheet"""
        ws = self.workbook.create_sheet("Statistical Analysis")
        
        if not measurements_queryset.exists():
            ws['A1'] = "No data available for analysis"
            return
        
        # Field mappings for analysis
        measurement_fields = {
            'Height at Withers': 'hauteur_au_garrot',
            'Body Length': 'body_length',
            'Chest Circumference': 'tour_de_poitrine',
            'Height at Croup': 'hauteur_au_sacrum',
            'Chest Width': 'largeur_poitrine',
            'Hip Width': 'largeur_hanche',
        }
        
        # Create analysis table
        ws['A1'] = "Measurement Statistics"
        ws['A1'].font = Font(size=14, bold=True)
        
        headers = ['Measurement', 'Count', 'Average', 'Min', 'Max', 'Std Dev']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
        
        row = 4
        for display_name, field_name in measurement_fields.items():
            # Filter out None values and calculate statistics
            values = [getattr(m, field_name) for m in measurements_queryset if getattr(m, field_name) is not None]
            
            if values:
                import statistics
                ws.cell(row=row, column=1, value=display_name)
                ws.cell(row=row, column=2, value=len(values))
                ws.cell(row=row, column=3, value=round(statistics.mean(values), 2))
                ws.cell(row=row, column=4, value=round(min(values), 2))
                ws.cell(row=row, column=5, value=round(max(values), 2))
                ws.cell(row=row, column=6, value=round(statistics.stdev(values), 2) if len(values) > 1 else 'N/A')
                row += 1
